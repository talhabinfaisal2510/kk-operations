from flask import Flask, render_template, request
import openpyxl
import random
import string
import os

app = Flask(__name__)

# Path to the Excel file
EXCEL_FILE_PATH = os.path.join(os.path.dirname(__file__), 'Book1.xlsx')

# Helper function to generate a unique ID
def generate_unique_id():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))

@app.route("/", methods=["GET"])
def home():
    return render_template("index.html")

@app.route("/submit", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        try:    
        # Extract form data
            date = request.form["date"]
            name = request.form["name"]
            referrer_contact = request.form["referrer_contact"]
            case_from = request.form["case_from"]
            family_member_contact = request.form["family_member_contact"]
            family_member_name = request.form["family_member_name"]
            address1 = request.form["address1"]
            case_detail = request.form["case_detail"]

            # Open the Excel file
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
                sheet = wb.active
            except PermissionError:
                return "Permission denied. Please check file permissions or if the file is open in another program."

            # Generate a unique ID for the new row
            unique_id = generate_unique_id()

            # Insert the new data into the Excel sheet
            new_row = [
                unique_id, date, name, referrer_contact, case_from, family_member_contact,family_member_name,
                address1, case_detail
            ]
            sheet.append(new_row)

            # Save the Excel file
            wb.save(EXCEL_FILE_PATH)

            return render_template("id_display.html", unique_id=unique_id)
        except Exception as e:
            # Log the error and show a user-friendly message
            print(f"Error in /submit: {str(e)}")
            return f"An error occurred while processing your request. Please try again later.Error :{str(e)}"
    return render_template("index.html")
@app.route("/referred_case_status", methods=["GET", "POST"])
def referred_case_status():
    case_data = None
    error_message = None
    
    if request.method == "POST":
        # Get the ID entered by the user
        case_id = request.form["case_id"]
        
        try:
            # Open the Excel workbook using openpyxl
            wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
            sheet = wb.active  # Assuming data is in the active sheet

            # Search for the ID in the first column (Unique ID column)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):  # Retrieve cell values directly
                if row[0] == case_id:  # Assuming the Unique ID is in the first column (A)
                    case_data = row
                    break

            # If no matching ID is found
            if not case_data:
                error_message = "No case found with this ID. Please check the ID and try again."
                
        except Exception as e:
            error_message = f"Error reading Excel file: {str(e)}"
        finally:
            wb.close()  # Ensure the workbook is closed after processing
        
    # Render the referred_case_status page with case_data (row) or error_message
    return render_template("referred_case_status.html", case_data=case_data, error_message=error_message)
    

if __name__ == "__main__":
    app.run(debug=True)
